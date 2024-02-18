import os
from PyQt5.QtWidgets import QWidget, QVBoxLayout, QHBoxLayout, QPushButton, QLabel, QSlider, QSizePolicy, QComboBox, QLineEdit, QScrollArea, QTableWidget, QTableWidgetItem, QPlainTextEdit
from PyQt5.QtCore import Qt, QSize, pyqtSignal
from PyQt5.QtGui import QPixmap, QImage
from natsort import natsorted
import pandas as pd
from stages.dataExtraction import execute_main
from PIL import Image
import chardet
from openpyxl import load_workbook

class DataExtractionWindow(QWidget):
    startSignal = pyqtSignal(str)  # Sygnał emitowany przy kliknięciu przycisku Start

    def __init__(self, parent=None):
        super().__init__(parent)
        self.parent = parent  # referencja do głównego okna

        self.main_layout = QVBoxLayout()
        self.buttons_layout = QHBoxLayout()
        self.csv_name_label = QLabel()
        self.csv_name_label.setAlignment(Qt.AlignCenter)
        
        self.class_input = QLineEdit()
        self.class_input.setPlaceholderText('Klasa zdjęcia')
        self.rename_button = QPushButton('Zmień typ')

        self.table_widget = QTableWidget()
        self.scroll_area = QScrollArea()
        self.scroll_area.setWidget(self.table_widget)
        self.scroll_area.setWidgetResizable(True)
        self.table_widget.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.csv_text_edit = QPlainTextEdit()

        self.start_button = QPushButton('Start')
        self.back_button = QPushButton('Cofnij')
        
        
        self.export_csv_button = QPushButton('Zapisz poprawione dane')

        self.image_name_label = QLabel()
        self.image_name_label.setAlignment(Qt.AlignCenter)

        self.image_label = QLabel()
        self.image_label.setAlignment(Qt.AlignCenter)
        self.image_label.setFixedSize(QSize(800, 400))
        self.image_label.setScaledContents(True)
        
        self.slider = QSlider(Qt.Horizontal)
        self.slider.setTickPosition(QSlider.TicksBelow)
        self.slider.setTickInterval(1)
        self.slider.setMinimum(1)
        self.merge_cells_button = QPushButton('Scal zaznaczone komórki')
        
        self.main_layout.addWidget(self.start_button)
        self.buttons_layout.addWidget(self.back_button)
        self.main_layout.addLayout(self.buttons_layout)

        self.main_layout.addWidget(self.image_label)  # Add image_label to the layout
        self.main_layout.addWidget(self.scroll_area)  # Add scroll_area to the layout
        self.main_layout.addWidget(self.csv_name_label)  # Add csv_name_label to the layout
        self.main_layout.addWidget(self.slider)
        self.main_layout.addWidget(self.merge_cells_button)
        self.main_layout.addWidget(self.export_csv_button)
        self.setLayout(self.main_layout)
        self.csv_files = self.getImageFiles()  # Get the list of image files
        self.merge_cells_button.clicked.connect(self.mergeSelectedCells)
        self.start_button.clicked.connect(self.startProcess)  # Connect the Start button to the startProcess function
        self.slider.valueChanged.connect(self.updateImage)  # Connect the slider value change to the updateImage function
        self.back_button.clicked.connect(self.goBack)
        self.export_csv_button.clicked.connect(self.exportToCSV)  # Connect the Export to CSV button to the exportToCSV function


        

    def exportToCSV(self):
        csv_index = self.slider.value()
        if self.csv_files:
            if csv_index < len(self.csv_files):
                csv_path = os.path.join('csv_classification', self.csv_files[csv_index])
                dataframe = pd.read_excel(csv_path, engine='openpyxl', header=None)
                print(self.table_widget.rowCount(),self.table_widget.columnCount())
                print(dataframe.shape)
                

                # Update the dataframe with the content of the table widget
                for row in range(self.table_widget.rowCount()):
                    for col in range(self.table_widget.columnCount()):
                        item = self.table_widget.item(row, col)
                        if item:
                            dataframe.iloc[row, col] = item.text()

                # Export the dataframe to a CSV file
                dataframe.to_excel(csv_path, index=False, header=False)
        

    def getImageFiles(self):
        csv_folder = 'csv_classification'  # Folder with csv files
        csv_files = natsorted([file for file in os.listdir(csv_folder) if file.endswith('.xlsx')])
        if csv_files:
            self.slider.setMaximum(len(csv_files) - 1)  # -1 because index starts from 0

            return csv_files

    def startProcess(self):
        # Perform processing with the selected option and coefficient
        execute_main()
        
        self.csv_files = self.getImageFiles()

        
    def resizeEvent(self, event):
        super().resizeEvent(event)
        self.updateImage()
        
    def mergeSelectedCells(self):
        selected_ranges = self.table_widget.selectedRanges()
        csv_index = self.slider.value()
        if self.csv_files:
            if csv_index < len(self.csv_files):
                csv_path = os.path.join('csv_classification', self.csv_files[csv_index])
                wb = load_workbook(filename=csv_path)
                ws = wb.active

                # Store all merged cells in a list
                all_merged_cells = []

                # Append existing merged cells to the list
                for merged_cells in ws.merged_cells.ranges:
                    all_merged_cells.append((merged_cells.min_row, merged_cells.min_col, merged_cells.max_row, merged_cells.max_col))

                top_rows = []
                bottom_rows = []
                left_columns = []
                right_columns = []

                # Append new selected ranges to the list
                for selected_range in selected_ranges:
                    top_row = selected_range.topRow()
                    bottom_row = selected_range.bottomRow()
                    left_column = selected_range.leftColumn()
                    right_column = selected_range.rightColumn()

                    top_rows.append(top_row)
                    bottom_rows.append(bottom_row)
                    left_columns.append(left_column)
                    right_columns.append(right_column)

                # Check if selected range is more than one cell
                if (max(bottom_rows) - min(top_rows) > 0) or (max(right_columns) - min(left_columns) > 0):
                    self.table_widget.setSpan(min(top_rows), min(left_columns), max(bottom_rows) - min(top_rows) + 1, max(right_columns) - min(left_columns) + 1)
                    all_merged_cells.append((min(top_rows) + 1, min(left_columns) + 1, max(bottom_rows) + 1, max(right_columns) + 1))



                # Merge cells according to the list
                for cell_range in all_merged_cells:
                    cell_values = []
                    for row in ws.iter_rows(min_row=cell_range[0], min_col=cell_range[1], max_row=cell_range[2], max_col=cell_range[3]):
                        for cell in row:
                            if cell.value:
                                cell_values.append(str(cell.value))
                    print(cell_values)
                    ws.merge_cells(start_row=cell_range[0], start_column=cell_range[1], end_row=cell_range[2], end_column=cell_range[3])
                    if cell_range == all_merged_cells[-1]:
                        ws.cell(row=cell_range[0], column=cell_range[1]).value = " ".join(cell_values)


                # Save the workbook, this will overwrite the existing file
                wb.save(csv_path)

                # Update the image to reflect the changes
                self.updateImage()

        
        
        
    def updateImage(self):
        csv_index = self.slider.value()
        if self.csv_files:
            if csv_index < len(self.csv_files):
                csv_path = os.path.join('csv_classification', self.csv_files[csv_index])
                self.csv_name_label.setText(self.csv_files[csv_index])
                print(csv_path)

                image_number = self.csv_files[csv_index].split('.')[0]
                image_path = f"images_classification/table_{image_number}.jpg"
                print(image_path)
                # Load and display the image
                image = QImage(image_path)

                # Scale the image to the width of the window
                window_width = self.width()
                scaled_image = image.scaledToWidth(window_width - 20)

                self.image_label.setPixmap(QPixmap.fromImage(scaled_image))

                wb = load_workbook(filename=csv_path)
                ws = wb.active

                # Remove the current QTableWidget from the layout and initialize a new one
                self.scroll_area.takeWidget()
                self.table_widget = QTableWidget()
                self.scroll_area.setWidget(self.table_widget)
                self.scroll_area.setWidgetResizable(True)
                self.table_widget.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

                # Update the table widget with the content from the active sheet
                self.table_widget.setRowCount(ws.max_row)
                self.table_widget.setColumnCount(ws.max_column)

                for row in ws.iter_rows(min_row=1, max_col=ws.max_column, max_row=ws.max_row):
                    for cell in row:
                        # Add data to the table widget
                        item_value = "" if cell.value is None else str(cell.value)
                        item = QTableWidgetItem(item_value)
                        self.table_widget.setItem(cell.row-1, cell.column-1, item)

                        # Merge cells in table widget if they are merged in the Excel sheet
                        if cell.coordinate in ws.merged_cells:
                            merged_cells = [c for c in ws.merged_cells.ranges if cell.coordinate in c.coord]
                            if merged_cells:  # Check if the list is not empty
                                merged_cell = merged_cells[0]
                                self.table_widget.setSpan(cell.row-1, cell.column-1, merged_cell.size['rows'], merged_cell.size['columns'])



    def goBack(self):
        print(self.parent)
        # Code to go back to the previous state or screen
        self.parent.stackedWidget.setCurrentWidget(self.parent.classificationWidget)
        self.parent.setWindowTitle('Classification')
   
            

if __name__ == '__main__':
    app = QApplication([])
    window = PreprocessingWindow()
    window.show()
    app.exec()